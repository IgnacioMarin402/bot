"""
EXPORTACIÓN a Excel de los datos que administra Julieta.

Daniela vive en Excel: este módulo convierte el almacén SQLite de vuelta a
su formato natural — un .xlsx con una hoja por categoría (ventas,
pendientes, portabilidades, homepass).

Uso (lo corre el dueño del proyecto, no el bot):

    uv run poe exportar

Genera exports/julieta_<fecha>_<hora>.xlsx y deja el archivo listo para
mandárselo. Es un script y no una tool a propósito: por WhatsApp el bot solo
puede responder texto (mandar archivos requiere hostearlos en una URL
pública — anotado como mejora futura en el roadmap).
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from bots.julieta import almacen

CARPETA_EXPORTS = Path(__file__).resolve().parents[2] / "exports"


def exportar_excel(ruta: Path | None = None) -> Path:
    """Genera el .xlsx con una hoja por categoría y devuelve su ruta."""
    if ruta is None:
        marca = datetime.now().strftime("%Y-%m-%d_%H%M")
        CARPETA_EXPORTS.mkdir(exist_ok=True)
        ruta = CARPETA_EXPORTS / f"julieta_{marca}.xlsx"

    libro = Workbook()
    libro.remove(libro.active)  # partir sin la hoja vacía por defecto

    for categoria in almacen.CATEGORIAS:
        hoja = libro.create_sheet(title=categoria)
        # listar() ordena del más reciente al más viejo; en el Excel se lee
        # mejor cronológico (como ella lo llenaría a mano) -> se invierte.
        filas = list(reversed(almacen.listar(categoria, limite=1_000_000)))
        if not filas:
            hoja.append(["(sin registros)"])
            continue
        encabezados = list(filas[0].keys())
        hoja.append(encabezados)
        for celda in hoja[1]:
            celda.font = Font(bold=True)
        for fila in filas:
            hoja.append([fila[col] for col in encabezados])
        # Ancho de columnas aproximado al contenido (legibilidad).
        for indice, col in enumerate(encabezados, start=1):
            largo = max(len(str(col)), *(len(str(f[col])) for f in filas))
            hoja.column_dimensions[hoja.cell(row=1, column=indice).column_letter].width = min(
                largo + 2, 45
            )

    libro.save(ruta)
    return ruta


if __name__ == "__main__":
    print("Excel generado:", exportar_excel())
